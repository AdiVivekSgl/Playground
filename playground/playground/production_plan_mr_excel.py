# Copyright (c) 2026, Frontec and contributors
# For license information, please see license.txt

"""
FGSRM Production Plan Excel (3 sheets)
======================================

A Playground-owned workbook the FGSRM "Create Prodn Plan" button downloads after
building a Production Plan - separate from frontec's own MR Hierarchy Excel
(`frontec.doc_event.production_plan.hierarchy_excel`), which is left untouched.

Sheets (matching the supplied template):

  1. "FG Plan"                - the FGSRM report view that produced the plan
                                (Item Name / Customer / SO / Dispatch Priority /
                                Suggested Prodn), rebuilt from the same filters.
  2. "FG Items Consolidated"  - the root plan's Assembly Items (po_items):
                                Item Code / BOM No / Planned Qty / UOM. Same as
                                frontec's "Assembly Items" sheet minus Warehouse
                                and Planned Start Date.
  3. "RM Plan"                - raw-material shortage across the plan's nested
                                chain. Same data/logic as frontec's "Combined"
                                sheet with columns removed and reordered; column
                                E "WO Qty" is an in-cell Excel FORMULA
                                =MAX(0, QtyAsPerBOM - QtyInStock) - PlanToRequest,
                                so it recalculates if the user edits the sheet.
                                "Short QTY" stays a static computed value
                                (= max(0, qty - in-stock - pending PO)), as in
                                frontec.

Built with openpyxl (bundled with Frappe) and streamed via frappe.response.
"""

from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt

from playground.playground.report.fg_stock_reservation_manager.fg_stock_reservation_manager import (
	execute as fgsrm_execute,
)

# frontec's parent-plan link that chains the nested Production Plans.
PARENT_FIELD = "custom_parent_production_plan"
MAX_LEVELS = 5

_HEADER_FILL = "D3D3D3"


@frappe.whitelist()
def download_fgsrm_mr_excel(name, filters=None):
	"""Stream the 3-sheet workbook for Production Plan `name`. `filters` is the
	FGSRM filter JSON (for the "FG Plan" sheet); optional so the endpoint still
	works when called with just a plan name."""
	if not frappe.has_permission("Production Plan", "read", doc=name):
		frappe.throw(
			_("You are not permitted to read Production Plan {0}.").format(name),
			frappe.PermissionError,
		)

	filters = frappe.parse_json(filters) if filters else {}

	import openpyxl

	wb = openpyxl.Workbook()
	wb.remove(wb.active)  # drop the default empty sheet

	_build_fg_plan_sheet(wb, filters)
	_build_fg_items_consolidated_sheet(wb, name)
	_build_rm_plan_sheet(wb, name)

	stream = BytesIO()
	wb.save(stream)

	frappe.response["filename"] = "MR_Hierarchy_{0}.xlsx".format(str(name).replace("/", "-"))
	frappe.response["filecontent"] = stream.getvalue()
	frappe.response["type"] = "binary"


# --------------------------------------------------------------------------- #
# Shared styling
# --------------------------------------------------------------------------- #

def _write_header(ws, headers, row=1):
	"""Grey, bold, centred header row (matches frontec's sheets)."""
	from openpyxl.styles import Alignment, Font, PatternFill

	fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
	font = Font(bold=True)
	center = Alignment(horizontal="center")
	for col, label in enumerate(headers, start=1):
		c = ws.cell(row, col, label)
		c.font = font
		c.fill = fill
		c.alignment = center


def _autosize(ws, headers):
	from openpyxl.utils import get_column_letter

	for col, label in enumerate(headers, start=1):
		ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(label)) + 2)


# --------------------------------------------------------------------------- #
# Sheet 1 - FG Plan (from the FGSRM report view)
# --------------------------------------------------------------------------- #

def _build_fg_plan_sheet(wb, filters):
	ws = wb.create_sheet("FG Plan")
	headers = ["Item Name", "Customer", "SO", "Dispatch Priority", "Suggested Prodn"]
	_write_header(ws, headers)

	# Per-line FGSRM view for the same filters (force the collapse off).
	view_filters = dict(filters or {})
	view_filters["group_by_so"] = 0
	_columns, data = fgsrm_execute(view_filters)

	for row in data:
		ws.append([
			row.get("item_name"),
			row.get("customer"),
			row.get("sales_order"),
			row.get("so_date"),
			flt(row.get("suggested_prodn")),
		])
	_autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 2 - FG Items Consolidated (root plan po_items)
# --------------------------------------------------------------------------- #

def _build_fg_items_consolidated_sheet(wb, plan_name):
	ws = wb.create_sheet("FG Items Consolidated")
	headers = ["Item Code", "BOM No", "Planned Qty", "UOM"]
	_write_header(ws, headers)

	root_name = _build_chain(plan_name)[0]
	po_items = frappe.get_all(
		"Production Plan Item",
		filters={"parent": root_name},
		fields=["item_code", "bom_no", "planned_qty", "stock_uom"],
		order_by="idx asc",
	)
	for row in po_items:
		ws.append([
			row.get("item_code"),
			row.get("bom_no"),
			flt(row.get("planned_qty")),
			row.get("stock_uom"),
		])
	_autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 3 - RM Plan (nested-chain raw-material shortage)
# --------------------------------------------------------------------------- #

def _build_rm_plan_sheet(wb, plan_name):
	ws = wb.create_sheet("RM Plan")
	headers = [
		"Item Code",            # A
		"Type",                 # B
		"Explosion Lvl",        # C
		"Qty As Per BOM",       # D
		"WO Qty",               # E  (in-cell formula, below)
		"Plan to Request Qty",  # F
		"Safety Stock",         # G
		"Minimum Order Qty",    # H
		"Qty In Stock",         # I
		"Ordered Qty",          # J
		"Short QTY",            # K  (static computed, as in frontec)
	]
	_write_header(ws, headers)

	rows = _collect_mr_rows(_build_chain(plan_name))
	if not rows:
		_autosize(ws, headers)
		return

	item_codes = sorted({r["item_code"] for r in rows})
	po_pending = _pending_po_map(item_codes)      # J: outstanding PO qty
	min_oqty = _item_field_map(item_codes, "min_order_qty")
	safety = _item_field_map(item_codes, "safety_stock")

	r = 2  # data starts at row 2 (headers on row 1)
	for row in rows:
		item = row["item_code"]
		qty = flt(row.get("quantity"))                 # Plan to Request Qty
		bom_qty = flt(row.get("required_bom_qty"))     # Qty As Per BOM
		actual_qty = flt(row.get("actual_qty"))        # Qty In Stock
		pending_po = flt(po_pending.get(item))         # Ordered Qty
		short_qty = max(0.0, qty - actual_qty - pending_po)

		ws.cell(r, 1, item)
		ws.cell(r, 2, row.get("material_request_type"))
		ws.cell(r, 3, row["_level"])
		ws.cell(r, 4, bom_qty)
		# E "WO Qty" - live formula, per the template: =MAX(0, D - I) - F
		ws.cell(r, 5, "=MAX(0,D{r}-I{r})-F{r}".format(r=r))
		ws.cell(r, 6, qty)
		ws.cell(r, 7, flt(safety.get(item)))
		ws.cell(r, 8, flt(min_oqty.get(item)))
		ws.cell(r, 9, actual_qty)
		ws.cell(r, 10, pending_po)
		ws.cell(r, 11, short_qty)
		r += 1

	_autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Data helpers
# --------------------------------------------------------------------------- #

def _build_chain(name):
	"""[root, ..., leaf] for the nested Production Plan chain (mirrors frontec's
	_build_hierarchy_chain): walk up via PARENT_FIELD to the root, then down the
	single-child chain to the leaf, capped at MAX_LEVELS against circular refs.
	Degrades to [name] where the frontec parent field isn't installed."""
	if not frappe.db.has_column("Production Plan", PARENT_FIELD):
		return [name]

	# Phase 1 - find root
	visited = set()
	current = name
	while current and len(visited) < MAX_LEVELS:
		if current in visited:
			break
		visited.add(current)
		parent = frappe.db.get_value("Production Plan", current, PARENT_FIELD)
		if not parent:
			break
		current = parent
	root = current

	# Phase 2 - walk down root -> leaf
	chain = []
	seen = set()
	current = root
	while current and len(chain) < MAX_LEVELS:
		if current in seen:
			break
		seen.add(current)
		chain.append(current)
		current = frappe.db.get_value("Production Plan", {PARENT_FIELD: current}, "name")
	return chain


def _collect_mr_rows(chain):
	"""Every Material Request Plan Item across the chain, tagged with its level
	(L1..Ln). Same fields frontec reads (minus the ones the template dropped)."""
	rows = []
	for idx, pp in enumerate(chain):
		level = "L{0}".format(idx + 1)
		for d in frappe.get_all(
			"Material Request Plan Item",
			filters={"parent": pp},
			fields=["item_code", "material_request_type", "quantity", "required_bom_qty", "actual_qty"],
			order_by="idx asc",
		):
			d["_level"] = level
			rows.append(d)
	return rows


def _pending_po_map(items):
	"""{item_code: outstanding PO qty} = Σ max(qty - received_qty, 0) over
	submitted, non-Closed/Cancelled Purchase Orders (frontec's po_pending)."""
	if not items:
		return {}
	rows = frappe.db.sql(
		"""
		SELECT poi.item_code,
			SUM(GREATEST(poi.qty - IFNULL(poi.received_qty, 0), 0)) AS pending_qty
		FROM `tabPurchase Order Item` poi
		INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
		WHERE poi.item_code IN %(codes)s
			AND po.docstatus = 1
			AND po.status NOT IN ('Closed', 'Cancelled')
		GROUP BY poi.item_code
		""",
		{"codes": items},
		as_dict=True,
	)
	return {r.item_code: flt(r.pending_qty) for r in rows}


def _item_field_map(items, field):
	"""{item_code: <field>} from the Item master (used for Minimum Order Qty and
	Safety Stock)."""
	if not items:
		return {}
	rows = frappe.db.sql(
		"SELECT name, `{0}` AS val FROM `tabItem` WHERE name IN %(codes)s".format(field),
		{"codes": items},
		as_dict=True,
	)
	return {r.name: flt(r.val) for r in rows}
