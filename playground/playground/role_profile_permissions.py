"""Role Profile permission workbook: export a Role Profile's roles, their full
permission grids, and a rolled-up effective view into a 3-sheet .xlsx; edit it;
re-upload to mirror it back onto ERPNext (with a preview + confirmation before any
deletion).

Wired to the core Role Profile form via `doctype_js` (see hooks.py) and
playground/public/js/role_profile_permissions.js. System Manager only.

NOTE ON SCOPE: a Role's permissions are global. Editing a role here changes it for
every profile and user that holds it — the workbook is organised by profile purely
for convenience. The client's confirmation dialog surfaces that blast radius.
"""

from io import BytesIO

import frappe
from frappe import _
from frappe.permissions import setup_custom_perms

# The 14 permission rights, in Role-Permissions-Manager order (after the If Owner /
# Level key columns). All are SQL-reserved or Python-keyword sensitive, so they are
# only ever read via backticked SQL and written via dict update — never bare attrs
# (`import` is a keyword; `read`/`create`/`delete`/`print`/`select` are SQL words).
RIGHTS = [
	"select", "read", "write", "create", "delete", "submit", "cancel",
	"amend", "report", "import", "export", "print", "email", "share",
]
RIGHT_LABELS = {
	"select": "Select", "read": "Read", "write": "Write", "create": "Create",
	"delete": "Delete", "submit": "Submit", "cancel": "Cancel", "amend": "Amend",
	"report": "Report", "import": "Import", "export": "Export", "print": "Print",
	"email": "Email", "share": "Share",
}

SHEET_ROLES = "Roles"
SHEET_PERMISSIONS = "Permissions"
SHEET_ROLLUP = "Roll-up"

ROLES_HEADERS = ["Role"]
PERM_HEADERS = ["Role", "Document Type", "Level", "If Owner"] + [RIGHT_LABELS[r] for r in RIGHTS]
ROLLUP_HEADERS = ["Document Type", "Level", "If Owner"] + [RIGHT_LABELS[r] for r in RIGHTS] + ["Granted By"]


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@frappe.whitelist()
def export_workbook(role_profile):
	"""Build and stream the 3-sheet workbook for a Role Profile. Sets a binary file
	response, so call it via open_url_post / a direct link, not frappe.call."""
	frappe.only_for("System Manager")
	if not frappe.db.exists("Role Profile", role_profile):
		frappe.throw(_("Role Profile {0} not found.").format(role_profile))

	import openpyxl
	from openpyxl.styles import Font

	roles = _profile_roles(role_profile)

	wb = openpyxl.Workbook()

	# Sheet 1 — Roles in the profile.
	ws_roles = wb.active
	ws_roles.title = SHEET_ROLES
	ws_roles.append(ROLES_HEADERS)
	for role in roles:
		ws_roles.append([role])

	# Sheet 2 — one row per Role x DocType x Level x If Owner, full flag set.
	ws_perms = wb.create_sheet(SHEET_PERMISSIONS)
	ws_perms.append(PERM_HEADERS)

	rollup = {}  # (doctype, level, if_owner) -> {right: 0/1, "_by": set(roles)}
	for role in roles:
		for r in sorted(_effective_perms(role), key=lambda x: (x["parent"], x["permlevel"], x["if_owner"])):
			level = int(r["permlevel"] or 0)
			if_owner = int(r["if_owner"] or 0)
			flags = {right: int(r.get(right) or 0) for right in RIGHTS}
			ws_perms.append([role, r["parent"], level, if_owner] + [flags[right] for right in RIGHTS])

			key = (r["parent"], level, if_owner)
			agg = rollup.setdefault(key, {right: 0 for right in RIGHTS})
			by = agg.setdefault("_by", set())
			granted = False
			for right in RIGHTS:
				if flags[right]:
					agg[right] = 1
					granted = True
			if granted:
				by.add(role)

	# Sheet 3 — rolled-up effective access with the contributing roles.
	ws_roll = wb.create_sheet(SHEET_ROLLUP)
	ws_roll.append(ROLLUP_HEADERS)
	for key in sorted(rollup.keys()):
		doctype, level, if_owner = key
		agg = rollup[key]
		ws_roll.append(
			[doctype, level, if_owner]
			+ [agg[right] for right in RIGHTS]
			+ [", ".join(sorted(agg.get("_by", set())))]
		)

	for sheet in (ws_roles, ws_perms, ws_roll):
		for cell in sheet[1]:
			cell.font = Font(bold=True)
		sheet.freeze_panes = "A2"

	buf = BytesIO()
	wb.save(buf)

	stamp = frappe.utils.now_datetime().strftime("%Y%m%d_%H%M")
	frappe.response["filename"] = f"RoleProfile_{frappe.scrub(role_profile)}_{stamp}.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"


# ---------------------------------------------------------------------------
# Import — preview (no writes) and apply (mirror)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def preview_import(role_profile, file_url):
	"""Parse the uploaded workbook and return the diff against live state. Writes
	nothing — the client renders this for confirmation."""
	frappe.only_for("System Manager")
	roles, perms = _read_workbook(file_url)
	return _build_diff(role_profile, roles, perms)


@frappe.whitelist()
def apply_import(role_profile, file_url):
	"""Re-parse the workbook and mirror it onto ERPNext: sync the profile's roles,
	then add/update/remove Custom DocPerm rows for the roles in the Permissions
	sheet. Only DocTypes that actually change are touched (each is converted to
	custom perms first, preserving existing standard perms), and the permission
	cache is cleared for them."""
	frappe.only_for("System Manager")
	roles, perms = _read_workbook(file_url)
	diff = _build_diff(role_profile, roles, perms)
	if not diff.get("ok"):
		frappe.throw("<br>".join(diff.get("errors", [])), title=_("Import Validation Failed"))

	# 1. Role Profile membership (exact mirror of the Roles sheet).
	if diff["membership"]["add"] or diff["membership"]["remove"]:
		profile = frappe.get_doc("Role Profile", role_profile)
		profile.set("roles", [])
		for role in roles:
			profile.append("roles", {"role": role})
		profile.save()  # Frappe core propagates the profile's roles to its users

	# 2. Permissions. Convert only the changing DocTypes to custom perms up front.
	affected = {item["doctype"] for item in diff["perm_add"] + diff["perm_update"] + diff["perm_remove"]}
	for dt in affected:
		setup_custom_perms(dt)

	for item in diff["perm_add"]:
		_upsert_custom_perm(item["doctype"], item["role"], item["level"], item["if_owner"], item["flags"])
	for item in diff["perm_update"]:
		_upsert_custom_perm(item["doctype"], item["role"], item["level"], item["if_owner"], item["flags"])
	for item in diff["perm_remove"]:
		_delete_custom_perm(item["doctype"], item["role"], item["level"], item["if_owner"])

	for dt in affected:
		frappe.clear_cache(doctype=dt)

	return {
		"success": True,
		"membership_added": len(diff["membership"]["add"]),
		"membership_removed": len(diff["membership"]["remove"]),
		"perms_added": len(diff["perm_add"]),
		"perms_updated": len(diff["perm_update"]),
		"perms_removed": len(diff["perm_remove"]),
	}


def _build_diff(role_profile, desired_roles, desired_perms):
	"""Compare the workbook against live state. Returns a structured diff, or
	{ok: False, errors: [...]} if the sheet references unknown roles/doctypes."""
	if not frappe.db.exists("Role Profile", role_profile):
		return {"ok": False, "errors": [_("Role Profile {0} not found.").format(frappe.utils.escape_html(role_profile))]}

	errors = []
	for role in sorted(set(desired_roles) | {k[0] for k in desired_perms}):
		if not frappe.db.exists("Role", role):
			errors.append(_("Role '{0}' does not exist.").format(frappe.utils.escape_html(role)))
	for dt in sorted({k[1] for k in desired_perms}):
		if not frappe.db.exists("DocType", dt):
			errors.append(_("DocType '{0}' does not exist.").format(frappe.utils.escape_html(dt)))
	if errors:
		return {"ok": False, "errors": errors}

	# Membership (exact mirror).
	current_roles = _profile_roles(role_profile)
	membership_add = [r for r in desired_roles if r not in current_roles]
	membership_remove = [r for r in current_roles if r not in desired_roles]

	# Permissions. Scope = roles named anywhere in the Permissions sheet, so a role
	# whose rows are all unticked still has its live perms removed. Only rows that
	# grant at least one flag count as "desired"; an all-zero row means "no perm".
	scope_roles = sorted({k[0] for k in desired_perms})
	desired_grant = {k: v for k, v in desired_perms.items() if any(v.values())}

	perm_add, perm_update, perm_remove = [], [], []
	for role in scope_roles:
		current = {
			(r["parent"], int(r["permlevel"] or 0), int(r["if_owner"] or 0)): {right: int(r.get(right) or 0) for right in RIGHTS}
			for r in _effective_perms(role)
		}
		desired = {(dt, lvl, ifo): flags for (rr, dt, lvl, ifo), flags in desired_grant.items() if rr == role}

		for (dt, lvl, ifo), flags in desired.items():
			key = (dt, lvl, ifo)
			if key not in current:
				perm_add.append({"role": role, "doctype": dt, "level": lvl, "if_owner": ifo, "flags": flags})
			elif current[key] != flags:
				perm_update.append({
					"role": role, "doctype": dt, "level": lvl, "if_owner": ifo, "flags": flags,
					"changes": [right for right in RIGHTS if current[key][right] != flags[right]],
				})
		for (dt, lvl, ifo) in current:
			if (dt, lvl, ifo) not in desired:
				perm_remove.append({"role": role, "doctype": dt, "level": lvl, "if_owner": ifo})

	# Global blast radius per affected role.
	affected_roles = sorted(set(scope_roles) | set(membership_add) | set(membership_remove))
	impact = {
		role: {
			"other_profiles": _other_profiles_count(role, role_profile),
			"users": _users_with_role_count(role),
		}
		for role in affected_roles
	}

	return {
		"ok": True,
		"role_profile": role_profile,
		"membership": {"add": membership_add, "remove": membership_remove},
		"perm_add": perm_add,
		"perm_update": perm_update,
		"perm_remove": perm_remove,
		"impact": impact,
	}


# ---------------------------------------------------------------------------
# Data access helpers
# ---------------------------------------------------------------------------

def _profile_roles(role_profile):
	doc = frappe.get_doc("Role Profile", role_profile)
	seen, out = set(), []
	for d in doc.roles:
		if d.role and d.role not in seen:
			seen.add(d.role)
			out.append(d.role)
	return out


def _effective_perms(role):
	"""[{parent, permlevel, if_owner, <rights...>}] — the effective permissions for
	one role, exactly as the Role Permissions Manager shows them: Custom DocPerm
	rows where a DocType has been customised, else its standard DocPerm rows."""
	cols = "parent, permlevel, if_owner, " + ", ".join(f"`{r}`" for r in RIGHTS)
	custom = frappe.db.sql(f"select {cols} from `tabCustom DocPerm` where role=%s", (role,), as_dict=True)
	standard = frappe.db.sql(f"select {cols} from `tabDocPerm` where role=%s", (role,), as_dict=True)

	# A DocType uses custom perms wholesale once any Custom DocPerm exists for it.
	custom_mode = set(frappe.db.sql_list("select distinct parent from `tabCustom DocPerm`"))

	rows = list(custom)
	rows.extend(r for r in standard if r["parent"] not in custom_mode)
	return rows


def _upsert_custom_perm(doctype, role, permlevel, if_owner, flags):
	name = frappe.db.get_value(
		"Custom DocPerm",
		{"parent": doctype, "role": role, "permlevel": permlevel, "if_owner": if_owner},
	)
	if name:
		perm = frappe.get_doc("Custom DocPerm", name)
	else:
		perm = frappe.get_doc({
			"doctype": "Custom DocPerm",
			"parent": doctype, "parenttype": "DocType", "parentfield": "permissions",
			"role": role, "permlevel": permlevel, "if_owner": if_owner,
		})
	perm.update({right: (1 if flags.get(right) else 0) for right in RIGHTS})
	perm.save(ignore_permissions=True)


def _delete_custom_perm(doctype, role, permlevel, if_owner):
	name = frappe.db.get_value(
		"Custom DocPerm",
		{"parent": doctype, "role": role, "permlevel": permlevel, "if_owner": if_owner},
	)
	if name:
		frappe.delete_doc("Custom DocPerm", name, ignore_permissions=True, force=True)


def _other_profiles_count(role, current_profile):
	parents = frappe.get_all("Has Role", filters={"parenttype": "Role Profile", "role": role}, distinct=True, pluck="parent")
	return len({p for p in parents if p != current_profile})


def _users_with_role_count(role):
	parents = frappe.get_all("Has Role", filters={"parenttype": "User", "role": role}, distinct=True, pluck="parent")
	return len(set(parents))


# ---------------------------------------------------------------------------
# Workbook reading
# ---------------------------------------------------------------------------

def _read_workbook(file_url):
	import openpyxl
	from frappe.utils.file_manager import get_file

	_name, content = get_file(file_url)
	wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
	return _read_roles_sheet(wb), _read_perms_sheet(wb)


def _read_roles_sheet(wb):
	if SHEET_ROLES not in wb.sheetnames:
		frappe.throw(_("The workbook has no '{0}' sheet.").format(SHEET_ROLES))
	ws = wb[SHEET_ROLES]
	seen, out = set(), []
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		if i == 0 or not row:
			continue
		val = _text(row[0]) if len(row) else ""
		if val and val not in seen:
			seen.add(val)
			out.append(val)
	return out


def _read_perms_sheet(wb):
	if SHEET_PERMISSIONS not in wb.sheetnames:
		frappe.throw(_("The workbook has no '{0}' sheet.").format(SHEET_PERMISSIONS))
	ws = wb[SHEET_PERMISSIONS]
	header = {}
	out = {}
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		if i == 0:
			header = {
				str(v).strip().lower(): idx
				for idx, v in enumerate(row or [])
				if v is not None and str(v).strip()
			}
			continue
		if not row:
			continue

		def cell(*names):
			for n in names:
				idx = header.get(n)
				if idx is not None and idx < len(row):
					return row[idx]
			return None

		role = _text(cell("role"))
		doctype = _text(cell("document type", "doctype", "doctype name", "doctype_name"))
		if not role or not doctype:
			continue
		level = _as_int(cell("level", "permlevel"))
		if_owner = _as_check(cell("if owner", "if_owner"))
		out[(role, doctype, level, if_owner)] = {
			right: _as_check(cell(RIGHT_LABELS[right].lower(), right)) for right in RIGHTS
		}
	return out


def _text(v):
	return str(v).strip() if v is not None else ""


def _as_int(v):
	try:
		return int(float(v)) if v not in (None, "") else 0
	except (TypeError, ValueError):
		return 0


def _as_check(v):
	if v is None:
		return 0
	if isinstance(v, (int, float)):
		return 1 if v else 0
	return 1 if str(v).strip().lower() in ("1", "true", "yes", "y", "x", "checked") else 0
