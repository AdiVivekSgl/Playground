# Copyright (c) 2026, Frontec and contributors
# For license information, please see license.txt

"""
Unified Planning Workbook (Cover + 6 sheets)
============================================

A single workbook the FGSRM report and the Weekly Planning Snapshot both download
after building a Production Plan. It walks the whole planning story end to end, in
one file, in one consistent format - and one sheet ("Approved for Purchase") is
drop-in ready for the Purchase Authorization Sheet's "Populate from Excel".

The Production Plan form keeps its own existing download (frontec's MR Hierarchy
Excel) untouched - this workbook is the FGSRM/WPS-side artifact only.

Sheets (the data sheets share one grammar: column A = Item Code, a grey/bold/
centred header on row 1, Float qty / Currency value, a bold TOTAL row; redundant
columns are kept out so each fact lives on exactly one sheet, and several columns
are in-cell formulas so the workbook stays live when edited):

  0. "Cover"                        - the report title (which differs by origin -
                                       "Shortage Against 4 Week Production" from
                                       FGSRM, "Urgent Shortage Against Weekly
                                       Commitment" from WPS), hyperlinked origin +
                                       plan, and a generated-on stamp.
  1. "FG Reservation Status"        - the FGSRM picture per open SO line: Pending,
                                       Reserved, Short to Complete (=Pending−Reserved,
                                       formula), Item Free Stock, Suggested Prodn,
                                       Material / Sales Status.
  2. "Production Requirement"       - per-SO-line production ask: Suggested,
                                       Committed, Valuation Rate, Committed Value
                                       (=Committed×Rate, formula).
  3. "Consolidated Requirement"     - per-item roll-up: Item Free Stock, Total
                                       Suggested, Committed Prodn (the plan's own
                                       root po_items, authoritative for both entry
                                       points, so it reconciles with sheets 4-6).
  4. "Item Requirement (BOM Levels)"- the nested plan chain exploded across every
                                       BOM level: Qty As Per BOM, live WO Qty
                                       formula, Plan to Request, stock/ordered,
                                       Short QTY.
  5. "Item Requirement (logic)"     - EXPERIMENTAL restatement of sheet 4 that makes
                                       each shortage's provenance explicit (rows
                                       reference sheet 4 by cell). Adds two fetched
                                       columns - Reserved against Open WO, Projected
                                       Incoming from Open WO - and a Shortage formula.
                                       Kept beside sheet 4 to validate the logic.
  6. "Approved for Purchase"        - the summary buy-list: purchasable items netted
                                       once against stock + pending POs. Column A = Item,
                                       column B = Qty exactly as the PAS reader expects
                                       (`purchase_authorization_sheet._read_approved_sheet`);
                                       columns C+ are enrichment (Vendor is picked up
                                       by the reader's header-aware column lookup).

Data sources by entry point:
  - From FGSRM: `filters` (the report's filter JSON) drives sheets 1-3; the plan
    `name` drives sheets 3-6. Committed at the line level mirrors Suggested (the
    plan was built from Suggested), and sheet 3 / the plan reconcile it item-wise.
  - From WPS:   `snapshot` (the Weekly Planning Snapshot name) drives sheets 1-3
    from the frozen lines incl. Committed Prodn and Buffer rows; the plan `name`
    drives sheets 3-6.

Built with openpyxl (bundled with Frappe) and streamed via frappe.response.
"""

from io import BytesIO

import frappe
from frappe import _
from frappe.utils import (
    cint,
    flt,
    format_datetime,
    formatdate,
    get_fullname,
    get_url,
    get_url_to_form,
    getdate,
    now_datetime,
)

from playground.playground.report.fg_stock_reservation_manager.fg_stock_reservation_manager import (
    execute as fgsrm_execute,
)
from playground.playground.report.production_requirement_report.production_requirement_report import (
    get_stock_map,
)

APPROVED_SHEET = "Approved for Purchase"  # must match purchase_authorization_sheet.APPROVED_SHEET
BOM_SHEET = "Item Requirement (BOM Levels)"  # the Logic sheet references this one by cell

# frontec's parent-plan link that chains the nested Production Plans (used to walk
# the plan chain for sheets 3-4).
PARENT_FIELD = "custom_parent_production_plan"
MAX_LEVELS = 5
_HEADER_FILL = "D3D3D3"


@frappe.whitelist()
def download_unified_planning_workbook(plan, filters=None, snapshot=None):
    """Stream the unified workbook for Production Plan `plan`: a Cover sheet
    followed by the four planning sheets.

    Exactly one of `filters` (FGSRM entry point) or `snapshot` (WPS entry point)
    is expected for sheets 1-2; sheets 2B/3/4 always come from the plan chain.
    Both are optional so the endpoint degrades gracefully (a missing source just
    yields an empty sheet 1/2 rather than an error)."""
    if not frappe.has_permission("Production Plan", "read", doc=plan):
        frappe.throw(
            _("You are not permitted to read Production Plan {0}.").format(plan),
            frappe.PermissionError,
        )
    if snapshot and not frappe.has_permission("Weekly Planning Snapshot", "read", doc=snapshot):
        frappe.throw(
            _("You are not permitted to read Weekly Planning Snapshot {0}.").format(snapshot),
            frappe.PermissionError,
        )

    filters = frappe.parse_json(filters) if filters else {}

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    lines = _planning_lines(filters, snapshot)
    committed_by_item = _plan_committed_by_item(plan)
    # Built once and shared, so the BOM-Levels, Logic and Approved sheets all
    # sit on the same rows in the same order (the Logic sheet references the
    # BOM-Levels sheet by row number).
    mr_rows = _collect_mr_rows(_build_chain(plan))

    _build_cover_sheet(wb, plan, filters, snapshot)
    _build_fg_status_sheet(wb, lines)
    _build_production_requirement_sheet(wb, lines)
    _build_consolidated_requirement_sheet(wb, lines, committed_by_item)
    _build_bom_levels_sheet(wb, mr_rows)
    _build_logic_sheet(wb, mr_rows)
    _build_approved_for_purchase_sheet(wb, mr_rows)

    stream = BytesIO()
    wb.save(stream)

    frappe.response["filename"] = "Unified_Planning_{0}.xlsx".format(str(plan).replace("/", "-"))
    frappe.response["filecontent"] = stream.getvalue()
    frappe.response["type"] = "binary"


# --------------------------------------------------------------------------- #
# Cover sheet
# --------------------------------------------------------------------------- #

# Human-readable labels for the FGSRM filters listed on the cover; anything not
# here falls back to a title-cased fieldname.
_FILTER_LABELS = {
    "item_code": "Item",
    "customer": "Customer",
    "date_basis": "Date Basis",
    "unreserved_basis": "Unreserved Stock Basis",
    "view_mode": "View",
    "only_unreserved": "Only Lines With Unreserved Pending",
    "group_by_so": "Group by Sales Order",
    "include_draft": "Include Draft SOs",
    "from_date": "From Date",
    "to_date": "To Date",
}


def _build_cover_sheet(wb, plan, filters, snapshot):
    """First sheet: the report title (which differs by origin), a hyperlinked
    reference to the origin document and the Production Plan, and a generated-on
    timestamp.

      - From WPS   -> "Urgent Shortage Against Weekly Commitment", origin = the
                      Weekly Planning Snapshot (with its snapshot date).
      - From FGSRM -> "Shortage Against 4 Week Production", origin = the FG Stock
                      Reservation Manager report view, with its filters listed.

    Origin Document and Production Plan cells hyperlink to their desk forms (the
    FGSRM origin links to the report); links resolve against the site's own URL."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Cover")
    is_wps = bool(snapshot)
    title = (
        _("Urgent Shortage Against Weekly Commitment")
        if is_wps
        else _("Shortage Against 4 Week Production")
    )

    # Title band, merged across the key/value columns.
    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    label_font = _bold()
    link_font = Font(color="0563C1", underline="single")

    def _kv(row, label, value, url=None):
        lc = ws.cell(row, 1, label)
        lc.font = label_font
        vc = ws.cell(row, 2, value)
        if url:
            vc.hyperlink = url
            vc.font = link_font

    r = 3
    if is_wps:
        _kv(r, _("Source"), _("Weekly Planning Snapshot")); r += 1
        _kv(r, _("Origin Document"), snapshot, get_url_to_form("Weekly Planning Snapshot", snapshot)); r += 1
        snap_date = frappe.db.get_value("Weekly Planning Snapshot", snapshot, "snapshot_date")
        if snap_date:
            _kv(r, _("Snapshot Date"), formatdate(snap_date)); r += 1
    else:
        _kv(r, _("Source"), _("FG Stock Reservation Manager")); r += 1
        _kv(
            r,
            _("Origin Document"),
            _("FG Stock Reservation Manager (report view)"),
            get_url("/app/query-report/FG Stock Reservation Manager"),
        ); r += 1

    _kv(r, _("Production Plan"), plan, get_url_to_form("Production Plan", plan)); r += 1
    _kv(r, _("Generated On"), format_datetime(now_datetime())); r += 1
    _kv(r, _("Prepared By"), get_fullname(frappe.session.user)); r += 1

    # FGSRM filters that produced the view.
    if not is_wps:
        r += 1
        ws.cell(r, 1, _("Filters Applied")).font = label_font
        r += 1
        applied = _humanize_filters(filters)
        if applied:
            for label, value in applied:
                _kv(r, label, value); r += 1
        else:
            ws.cell(r, 1, _("(none — full open-order view)")); r += 1

    # Contents index.
    r += 1
    ws.cell(r, 1, _("Contents")).font = label_font
    r += 1
    for name in (
        _("1. FG Reservation Status"),
        _("2. Production Requirement"),
        _("3. Consolidated Requirement"),
        _("4. Item Requirement (BOM Levels)"),
        _("5. Item Requirement (logic)"),
        _("6. Approved for Purchase"),
    ):
        ws.cell(r, 1, name); r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 52


def _humanize_filters(filters):
    """[(label, value), ...] for the non-empty FGSRM filters, so the cover shows
    exactly the view that produced the plan. Off/blank/zero values are omitted;
    checkbox-style truthy values render as 'Yes'; lists are comma-joined."""
    out = []
    for key, value in (filters or {}).items():
        if isinstance(value, (list, tuple)):
            if not value:
                continue
            shown = ", ".join(str(v) for v in value)
        elif value in (None, "", 0, "0", False):
            continue
        elif value in (1, "1", True):
            shown = _("Yes")
        else:
            shown = str(value)
        label = _(_FILTER_LABELS.get(key) or key.replace("_", " ").title())
        out.append((label, shown))
    return out


# --------------------------------------------------------------------------- #
# Shared line model for sheets 1 & 2
# --------------------------------------------------------------------------- #

def _planning_lines(filters, snapshot):
    """Normalise the demand into one list of per-line dicts both sheet 1 and
    sheet 2 render from, so the two never drift.

    Keys: item_code, item_name, customer, sales_order, so_date, pending_qty,
    reserved_qty, item_free_stock, suggested_prodn, committed_prodn,
    valuation_rate, pending_value, short_to_complete, total_reserved_qty,
    reserved_by_customer, material_status, sales_status, source, is_buffer."""
    if snapshot:
        return _lines_from_snapshot(snapshot)
    if filters:
        return _lines_from_fgsrm(filters)
    return []


def _lines_from_snapshot(snapshot):
    """Per-line model from a Weekly Planning Snapshot's frozen items (incl. Buffer
    rows). Committed Prodn and Valuation Rate come straight off the snapshot;
    Material/Sales Status are read live from the Sales Orders (the snapshot doesn't
    freeze them)."""
    doc = frappe.get_doc("Weekly Planning Snapshot", snapshot)
    sos = sorted({d.sales_order for d in doc.items if d.sales_order})
    material_status, sales_status = _so_status_maps(sos)

    lines = []
    for d in doc.items:
        pending = flt(d.pending_qty)
        reserved = flt(d.reserved_qty)
        lines.append(
            {
                "item_code": d.item_code,
                "item_name": d.item_name,
                "customer": d.customer,
                "sales_order": d.sales_order,
                "so_date": d.so_date,
                "pending_qty": pending,
                "reserved_qty": reserved,
                "item_free_stock": flt(d.item_free_stock),
                "suggested_prodn": flt(d.suggested_prodn),
                "committed_prodn": flt(d.committed_prodn),
                "valuation_rate": flt(d.valuation_rate),
                # A snapshot line carries no selling rate, so Pending Value is left
                # to sheet 1's blank; the sheet totals qty, not value, for WPS.
                "pending_value": 0.0,
                "short_to_complete": max(0.0, pending - reserved),
                "total_reserved_qty": None,
                "reserved_by_customer": "",
                "material_status": material_status.get(d.sales_order),
                "sales_status": sales_status.get(d.sales_order),
                "source": _("Buffer") if d.is_buffer else "",
                "is_buffer": cint(d.is_buffer),
            }
        )
    return lines


def _lines_from_fgsrm(filters):
    """Per-line model from the live FGSRM report for `filters`. Committed Prodn
    mirrors Suggested Prodn (the plan is built from Suggested); Valuation Rate is
    enriched from STOCK_WAREHOUSE so sheet 2 can still value the commitment. The
    report's own TOTAL row is dropped - each sheet appends its own."""
    _columns, rows = fgsrm_execute(dict(filters or {}))
    rows = [r for r in rows if not r.get("is_total")]

    item_codes = sorted({r.get("item_code") for r in rows if r.get("item_code")})
    stock_map = get_stock_map(item_codes)

    lines = []
    for r in rows:
        suggested = flt(r.get("suggested_prodn"))
        stock = stock_map.get(r.get("item_code")) or frappe._dict()
        lines.append(
            {
                "item_code": r.get("item_code"),
                "item_name": r.get("item_name"),
                "customer": r.get("customer"),
                "sales_order": r.get("sales_order"),
                "so_date": r.get("so_date"),
                "pending_qty": flt(r.get("pending_qty")),
                "reserved_qty": flt(r.get("reserved_qty")),
                "item_free_stock": flt(r.get("item_free_stock")),
                "suggested_prodn": suggested,
                "committed_prodn": suggested,
                "valuation_rate": flt(stock.get("valuation_rate")),
                "pending_value": flt(r.get("pending_value")),
                "short_to_complete": flt(r.get("short_to_complete")),
                "total_reserved_qty": r.get("total_reserved_qty"),
                "reserved_by_customer": r.get("reserved_by_customer") or "",
                "material_status": r.get("material_status"),
                "sales_status": r.get("sales_status"),
                "source": r.get("source") or "",
                "is_buffer": 0,
            }
        )
    return lines


def _so_status_maps(sos):
    """({so: material_status}, {so: sales_status}) read live from the Sales Orders,
    each column fetched only where it exists (so this runs on a site missing one)."""
    if not sos:
        return {}, {}
    fields = ["name"]
    has_material = frappe.db.has_column("Sales Order", "custom_material_status")
    has_sales = frappe.db.has_column("Sales Order", "custom_sales_status")
    if has_material:
        fields.append("custom_material_status")
    if has_sales:
        fields.append("custom_sales_status")
    material, sales = {}, {}
    if len(fields) > 1:
        for r in frappe.get_all("Sales Order", filters={"name": ["in", sos]}, fields=fields):
            material[r.name] = r.get("custom_material_status") if has_material else None
            sales[r.name] = r.get("custom_sales_status") if has_sales else None
    return material, sales


def _line_sort_key(line):
    """Dispatch Priority Date, then Customer; Buffer rows (no date) fall to the
    bottom - matching the WPS Detailed view ordering."""
    sd = line.get("so_date")
    return (
        getdate(sd) if sd else getdate("9999-12-31"),
        1 if line.get("is_buffer") else 0,
        line.get("customer") or "",
        line.get("item_code") or "",
    )


# --------------------------------------------------------------------------- #
# Sheet 1 - FG Reservation Status
# --------------------------------------------------------------------------- #

def _build_fg_status_sheet(wb, lines):
    ws = wb.create_sheet("FG Reservation Status")
    headers = [
        "Item Code",            # A
        "Customer",             # B
        "SO",                   # C
        "Dispatch Priority Date",  # D
        "Pending Qty",          # E
        "Reserved Qty",         # F
        "Short to Complete",    # G  (formula: Pending − Reserved)
        "Item Free Stock",      # H
        "Suggested Prodn",      # I
        "Material Status",      # J
        "Sales Status",         # K
        "Source",               # L
    ]
    _write_header(ws, headers)

    r = 2
    for line in sorted(lines, key=_line_sort_key):
        ws.cell(r, 1, line.get("item_code"))
        ws.cell(r, 2, line.get("customer"))
        ws.cell(r, 3, line.get("sales_order"))
        ws.cell(r, 4, line.get("so_date"))
        ws.cell(r, 5, flt(line.get("pending_qty")))
        ws.cell(r, 6, flt(line.get("reserved_qty")))
        ws.cell(r, 7, "=E{r}-F{r}".format(r=r))  # Short to Complete = Pending − Reserved
        ws.cell(r, 8, flt(line.get("item_free_stock")))
        ws.cell(r, 9, flt(line.get("suggested_prodn")))
        ws.cell(r, 10, line.get("material_status"))
        ws.cell(r, 11, line.get("sales_status"))
        ws.cell(r, 12, line.get("source"))
        r += 1

    # TOTAL row - sum only the columns that legitimately add up (Item Free Stock is
    # a per-item figure repeated on every line, so it is deliberately not summed).
    if lines:
        _bold_cells(ws, r, {
            1: _("TOTAL"),
            5: sum(flt(l.get("pending_qty")) for l in lines),
            6: sum(flt(l.get("reserved_qty")) for l in lines),
            9: sum(flt(l.get("suggested_prodn")) for l in lines),
        })
    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 2 - Production Requirement (per-SO line detail)
# --------------------------------------------------------------------------- #

def _build_production_requirement_sheet(wb, lines):
    ws = wb.create_sheet("Production Requirement")
    headers = [
        "Item Code",        # A
        "Customer",         # B
        "SO",               # C
        "Dispatch Priority Date",  # D
        "Suggested Prodn",  # E
        "Committed Prodn",  # F
        "Valuation Rate",   # G
        "Committed Value",  # H  (formula: Committed × Rate)
    ]
    _write_header(ws, headers)

    r = 2
    for line in sorted(lines, key=_line_sort_key):
        ws.cell(r, 1, line.get("item_code"))
        ws.cell(r, 2, line.get("customer"))
        ws.cell(r, 3, line.get("sales_order"))
        ws.cell(r, 4, line.get("so_date"))
        ws.cell(r, 5, flt(line.get("suggested_prodn")))
        ws.cell(r, 6, flt(line.get("committed_prodn")))
        ws.cell(r, 7, flt(line.get("valuation_rate")))
        ws.cell(r, 8, "=F{r}*G{r}".format(r=r))  # Committed Value = Committed × Rate
        r += 1

    if lines:
        last = r - 1
        _bold_cells(ws, r, {
            1: _("TOTAL"),
            5: sum(flt(l.get("suggested_prodn")) for l in lines),
            6: sum(flt(l.get("committed_prodn")) for l in lines),
            8: "=SUM(H2:H{last})".format(last=last),
        })
    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 3 - Consolidated Requirement (per item)
# --------------------------------------------------------------------------- #

def _build_consolidated_requirement_sheet(wb, lines, committed_by_item):
    """Per-item roll-up: Committed is the plan's own root po_items (authoritative
    for both entry points, so it reconciles with the BOM/purchase sheets); Item
    Free Stock and Total Suggested are folded from the per-line model."""
    ws = wb.create_sheet("Consolidated Requirement")
    headers = [
        "Item Code",        # A
        "Item Free Stock",  # B
        "Total Suggested",  # C
        "Committed Prodn",  # D
    ]
    _write_header(ws, headers)

    by_item = {}
    order = []
    for line in lines:
        ic = line.get("item_code")
        if ic not in by_item:
            # Item Free Stock is a per-item fact: take it once, don't sum it.
            by_item[ic] = {"item_free_stock": flt(line.get("item_free_stock")), "suggested": 0.0}
            order.append(ic)
        by_item[ic]["suggested"] += flt(line.get("suggested_prodn"))

    # Include any plan item that carried no demand line, so the summary is the
    # full committed set that seeded the plan.
    for ic in committed_by_item:
        if ic not in by_item:
            by_item[ic] = {"item_free_stock": 0.0, "suggested": 0.0}
            order.append(ic)

    r = 2
    tot_sug = tot_com = 0.0
    for ic in order:
        agg = by_item[ic]
        committed = flt(committed_by_item.get(ic))
        ws.cell(r, 1, ic)
        ws.cell(r, 2, flt(agg["item_free_stock"]))
        ws.cell(r, 3, flt(agg["suggested"]))
        ws.cell(r, 4, committed)
        tot_sug += flt(agg["suggested"])
        tot_com += committed
        r += 1
    if order:
        _bold_cells(ws, r, {1: _("TOTAL"), 3: tot_sug, 4: tot_com})
    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 4 - Item Requirement (BOM Levels)  [nested-chain raw-material shortage]
# --------------------------------------------------------------------------- #

def _build_bom_levels_sheet(wb, mr_rows):
    ws = wb.create_sheet(BOM_SHEET)
    headers = [
        "Item Code",            # A
        "Type",                 # B
        "Explosion Lvl",        # C
        "Qty As Per BOM",       # D
        "WO Qty",               # E  (in-cell formula, below)
        "Plan to Request Qty",  # F
        "Safety Stock",         # G
        "Minimum Order Qty",    # H
        "Qty In Stock",         # I
        "Ordered Qty",          # J
        "Short QTY",            # K  (static computed: max(0, qty - in-stock - pending PO))
    ]
    _write_header(ws, headers)

    if not mr_rows:
        _autosize(ws, headers)
        return

    item_codes = sorted({r["item_code"] for r in mr_rows})
    po_pending = _pending_po_map(item_codes)      # J: outstanding PO qty
    min_oqty = _item_field_map(item_codes, "min_order_qty")
    safety = _item_field_map(item_codes, "safety_stock")

    r = 2  # data starts at row 2 (headers on row 1)
    for row in mr_rows:
        item = row["item_code"]
        qty = flt(row.get("quantity"))                 # Plan to Request Qty
        bom_qty = flt(row.get("required_bom_qty"))     # Qty As Per BOM
        actual_qty = flt(row.get("actual_qty"))        # Qty In Stock
        pending_po = flt(po_pending.get(item))         # Ordered Qty
        short_qty = max(0.0, qty - actual_qty - pending_po)

        ws.cell(r, 1, item)
        ws.cell(r, 2, row.get("material_request_type"))
        ws.cell(r, 3, row["_level"])
        ws.cell(r, 4, bom_qty)
        # E "WO Qty" - live formula: =MIN(0, PlanToRequest + QtyInStock + OrderedQty − QtyAsPerBOM)
        ws.cell(r, 5, "=MIN(0,(F{r}+I{r}+J{r}-D{r}))".format(r=r))
        ws.cell(r, 6, qty)
        ws.cell(r, 7, flt(safety.get(item)))
        ws.cell(r, 8, flt(min_oqty.get(item)))
        ws.cell(r, 9, actual_qty)
        ws.cell(r, 10, pending_po)
        ws.cell(r, 11, short_qty)
        r += 1

    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 5 - Item Requirement (logic)  [experimental, kept beside the std sheet]
# --------------------------------------------------------------------------- #

def _build_logic_sheet(wb, mr_rows):
    """Experimental restatement of the BOM-level requirement that makes each
    shortage's provenance explicit. Its rows mirror the "Item Requirement (BOM
    Levels)" sheet one-for-one (same order) - Item / Type / Level / Qty As Per BOM
    / Qty In Stock / Ordered Qty are pulled from it by cell reference so the two
    stay linked - and it adds two fetched columns that show what a line is blocked
    against and what is already coming:

        Shortage = MAX(0, Qty As Per BOM − Qty In Stock − Reserved against Open WO
                          + Projected Incoming from Open WO + Ordered Qty)

    Kept side by side with the standard sheet for a few weeks to validate the
    logic before it (potentially) replaces it - so the exact definition of the two
    fetched columns is deliberately easy to change (see the two _*_open_wo_map
    helpers)."""
    ws = wb.create_sheet("Item Requirement (logic)")
    headers = [
        "Item Code",                        # A  (= BOM!A)
        "Type",                             # B  (= BOM!B)
        "Explosion Lvl",                    # C  (= BOM!C)
        "Qty As Per BOM",                   # D  (= BOM!D)
        "Qty In Stock",                     # E  (= BOM!I)
        "Reserved against Open WO",         # F  (fetched)
        "Projected Incoming from Open WO",  # G  (fetched)
        "Incoming Purchase (Ordered Qty)",  # H  (= BOM!J)
        "Shortage",                         # I  (formula)
    ]
    _write_header(ws, headers)

    if not mr_rows:
        _autosize(ws, headers)
        return

    item_codes = sorted({r["item_code"] for r in mr_rows})
    reserved_wo = _reserved_against_open_wo_map(item_codes)
    incoming_wo = _projected_incoming_from_open_wo_map(item_codes)

    bom = "'{0}'".format(BOM_SHEET)  # quoted sheet name for the cross-sheet refs
    r = 2
    for row in mr_rows:
        item = row["item_code"]
        ws.cell(r, 1, "={0}!A{1}".format(bom, r))
        ws.cell(r, 2, "={0}!B{1}".format(bom, r))
        ws.cell(r, 3, "={0}!C{1}".format(bom, r))
        ws.cell(r, 4, "={0}!D{1}".format(bom, r))
        ws.cell(r, 5, "={0}!I{1}".format(bom, r))
        ws.cell(r, 6, flt(reserved_wo.get(item)))
        ws.cell(r, 7, flt(incoming_wo.get(item)))
        ws.cell(r, 8, "={0}!J{1}".format(bom, r))
        ws.cell(r, 9, "=MAX(0,(D{r}-E{r}-F{r}+G{r}+H{r}))".format(r=r))
        r += 1

    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Sheet 6 - Approved for Purchase  [PAS upload: A = Item, B = Qty]
# --------------------------------------------------------------------------- #

def _build_approved_for_purchase_sheet(wb, mr_rows):
    ws = wb.create_sheet(APPROVED_SHEET)
    headers = [
        "Item",        # A  ← PAS reads item_code from here
        "Qty",         # B  ← PAS reads the purchase qty from here
        "UOM",         # C  ── column C on is enrichment; the PAS reader ignores /
        "Rate",        # D     re-derives it, except Vendor which it picks up by
        "Value",       # E  (formula: Rate × Qty)   the header-aware column lookup.
        "Vendor",      # F
        "Lead Time",   # G
    ]
    _write_header(ws, headers)

    # Purchase-type lines only; net each item ONCE against stock + pending POs
    # (an item can recur across levels, so summing per-row Short QTY would
    # over-subtract its stock - aggregate Plan to Request first, then net).
    purchase_items = sorted({r["item_code"] for r in mr_rows if r.get("material_request_type") == "Purchase"})
    if not purchase_items:
        _autosize(ws, headers)
        return

    plan_req = {}
    actual_by_item = {}
    for row in mr_rows:
        if row.get("material_request_type") != "Purchase":
            continue
        item = row["item_code"]
        plan_req[item] = plan_req.get(item, 0.0) + flt(row.get("quantity"))
        actual_by_item[item] = flt(row.get("actual_qty"))  # per-item constant

    po_pending = _pending_po_map(purchase_items)
    info = _item_info_map(purchase_items)
    vendors = _default_supplier_map(purchase_items)

    r = 2
    first = r
    tot_qty = 0.0
    for item in purchase_items:
        net = max(0.0, flt(plan_req.get(item)) - flt(actual_by_item.get(item)) - flt(po_pending.get(item)))
        if net <= 0:
            continue
        it = info.get(item) or frappe._dict()
        ws.cell(r, 1, item)
        ws.cell(r, 2, net)
        ws.cell(r, 3, it.get("stock_uom"))
        ws.cell(r, 4, flt(it.get("valuation_rate")))
        ws.cell(r, 5, "=D{r}*B{r}".format(r=r))  # Value = Rate × Qty
        ws.cell(r, 6, vendors.get(item))
        ws.cell(r, 7, cint(it.get("lead_time_days")))
        tot_qty += net
        r += 1

    # TOTAL row - the PAS reader skips a row whose column A is "Total", so this is
    # safe to leave in the upload.
    if r > first:
        _bold_cells(ws, r, {
            1: _("Total"),
            2: tot_qty,
            5: "=SUM(E{a}:E{b})".format(a=first, b=r - 1),
        })
    _autosize(ws, headers)


# --------------------------------------------------------------------------- #
# Data helpers
# --------------------------------------------------------------------------- #

def _plan_committed_by_item(plan_name):
    """{item_code: planned_qty} from the plan chain's ROOT po_items - the
    authoritative committed production that seeded the whole plan (the root
    Assembly Items of the nested Production Plan)."""
    root = _build_chain(plan_name)[0]
    out = {}
    for row in frappe.get_all(
        "Production Plan Item",
        filters={"parent": root},
        fields=["item_code", "planned_qty"],
        order_by="idx asc",
    ):
        out[row.item_code] = out.get(row.item_code, 0.0) + flt(row.planned_qty)
    return out


def _item_info_map(items):
    """{item_code: {item_name, stock_uom, valuation_rate, lead_time_days}} for the
    Approved for Purchase enrichment columns."""
    if not items:
        return {}
    return {
        r.name: r
        for r in frappe.get_all(
            "Item",
            filters={"name": ["in", items]},
            fields=["name", "item_name", "stock_uom", "valuation_rate", "lead_time_days"],
        )
    }


def _default_supplier_map(items):
    """{item_code: default_supplier} from Item Default (first row per item)."""
    if not items:
        return {}
    out = {}
    for r in frappe.get_all(
        "Item Default",
        filters={"parent": ["in", items], "default_supplier": ["is", "set"]},
        fields=["parent", "default_supplier"],
    ):
        out.setdefault(r.parent, r.default_supplier)
    return out


# Open Work Orders = submitted, not yet finished/stopped/cancelled.
_OPEN_WO_STATUSES = ("Not Started", "In Process")


def _reserved_against_open_wo_map(items):
    """{item_code: qty of this component still required by OPEN Work Orders} =
    Σ max(required_qty − transferred_qty, 0) over Work Order Items whose parent
    Work Order is open (submitted, status Not Started / In Process) - the stock
    already blocked against production in the pipeline.

    Feeds the experimental "Item Requirement (logic)" sheet's "Reserved against
    Open WO" column. Deliberately isolated so the definition is easy to change
    while the sheet is validated against the standard one - e.g. swap to Stock
    Reservation Entries (voucher_type='Work Order') if that becomes the agreed
    basis."""
    if not items:
        return {}
    rows = frappe.db.sql(
        """
        SELECT woi.item_code AS item_code,
            SUM(GREATEST(woi.required_qty - IFNULL(woi.transferred_qty, 0), 0)) AS qty
        FROM `tabWork Order Item` woi
        INNER JOIN `tabWork Order` wo ON wo.name = woi.parent
        WHERE wo.docstatus = 1
            AND wo.status IN %(statuses)s
            AND woi.item_code IN %(items)s
        GROUP BY woi.item_code
        """,
        {"items": items, "statuses": _OPEN_WO_STATUSES},
        as_dict=True,
    )
    return {r.item_code: flt(r.qty) for r in rows}


def _projected_incoming_from_open_wo_map(items):
    """{item_code: qty still to be produced by OPEN Work Orders} =
    Σ max(qty − produced_qty, 0) over open Work Orders (submitted, status Not
    Started / In Process) whose production_item is this item - the supply already
    in the production pipeline.

    Feeds the experimental "Item Requirement (logic)" sheet's "Projected Incoming
    from Open WO" column. See _reserved_against_open_wo_map on why this is kept
    separate and easy to adjust."""
    if not items:
        return {}
    rows = frappe.db.sql(
        """
        SELECT production_item AS item_code,
            SUM(GREATEST(qty - IFNULL(produced_qty, 0), 0)) AS qty
        FROM `tabWork Order`
        WHERE docstatus = 1
            AND status IN %(statuses)s
            AND production_item IN %(items)s
        GROUP BY production_item
        """,
        {"items": items, "statuses": _OPEN_WO_STATUSES},
        as_dict=True,
    )
    return {r.item_code: flt(r.qty) for r in rows}


# --------------------------------------------------------------------------- #
# Plan-chain helpers (walk the nested Production Plan for sheets 3-4)
# --------------------------------------------------------------------------- #

def _build_chain(name):
    """[root, ..., leaf] for the nested Production Plan chain (mirrors frontec's
    _build_hierarchy_chain): walk up via PARENT_FIELD to the root, then down the
    single-child chain to the leaf, capped at MAX_LEVELS against circular refs.
    Degrades to [name] where the frontec parent field isn't installed."""
    if not frappe.db.has_column("Production Plan", PARENT_FIELD):
        return [name]

    # Phase 1 - find root
    visited = set()
    current = name
    while current and len(visited) < MAX_LEVELS:
        if current in visited:
            break
        visited.add(current)
        parent = frappe.db.get_value("Production Plan", current, PARENT_FIELD)
        if not parent:
            break
        current = parent
    root = current

    # Phase 2 - walk down root -> leaf
    chain = []
    seen = set()
    current = root
    while current and len(chain) < MAX_LEVELS:
        if current in seen:
            break
        seen.add(current)
        chain.append(current)
        current = frappe.db.get_value("Production Plan", {PARENT_FIELD: current}, "name")
    return chain


def _collect_mr_rows(chain):
    """Every Material Request Plan Item across the chain, tagged with its level
    (L1..Ln)."""
    rows = []
    for idx, pp in enumerate(chain):
        level = "L{0}".format(idx + 1)
        for d in frappe.get_all(
            "Material Request Plan Item",
            filters={"parent": pp},
            fields=["item_code", "material_request_type", "quantity", "required_bom_qty", "actual_qty"],
            order_by="idx asc",
        ):
            d["_level"] = level
            rows.append(d)
    return rows


def _pending_po_map(items):
    """{item_code: outstanding PO qty} = Σ max(qty - received_qty, 0) over
    submitted, non-Closed/Cancelled Purchase Orders (frontec's po_pending)."""
    if not items:
        return {}
    rows = frappe.db.sql(
        """
        SELECT poi.item_code,
            SUM(GREATEST(poi.qty - IFNULL(poi.received_qty, 0), 0)) AS pending_qty
        FROM `tabPurchase Order Item` poi
        INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE poi.item_code IN %(codes)s
            AND po.docstatus = 1
            AND po.status NOT IN ('Closed', 'Cancelled')
        GROUP BY poi.item_code
        """,
        {"codes": items},
        as_dict=True,
    )
    return {r.item_code: flt(r.pending_qty) for r in rows}


def _item_field_map(items, field):
    """{item_code: <field>} from the Item master (used for Minimum Order Qty and
    Safety Stock)."""
    if not items:
        return {}
    rows = frappe.db.sql(
        "SELECT name, `{0}` AS val FROM `tabItem` WHERE name IN %(codes)s".format(field),
        {"codes": items},
        as_dict=True,
    )
    return {r.name: flt(r.val) for r in rows}


# --------------------------------------------------------------------------- #
# Small styling helpers
# --------------------------------------------------------------------------- #

def _write_header(ws, headers, row=1):
    """Grey, bold, centred header row."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
    font = Font(bold=True)
    center = Alignment(horizontal="center")
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row, col, label)
        c.font = font
        c.fill = fill
        c.alignment = center


def _autosize(ws, headers):
    from openpyxl.utils import get_column_letter

    for col, label in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(label)) + 2)


def _bold():
    from openpyxl.styles import Font

    return Font(bold=True)


def _bold_cells(ws, row, values):
    """Write {column_index: value} into `row`, each cell bold - used for the TOTAL
    rows. A string value starting with '=' is stored by openpyxl as a formula; a
    label string as text; a number as a number. Columns not listed stay blank."""
    bold = _bold()
    for col, val in values.items():
        c = ws.cell(row, col, val)
        c.font = bold
    return row
